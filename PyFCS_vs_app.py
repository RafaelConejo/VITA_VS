import os
import sys
import random
import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import ttk, simpledialog, messagebox
from PIL import Image, ImageTk
import pandas as pd
import ast

current_dir = os.path.dirname(__file__)
sys.path.append(current_dir)


############################################################ FUNTIONS ############################################################

def exit_fullscreen(event):
    root.attributes('-fullscreen', False)


def parse_column(cell):
    try:
        # Usar `ast.literal_eval` para evaluar la cadena de forma segura
        return ast.literal_eval(cell) if isinstance(cell, str) else cell
    except Exception as e:
        print(f"Error al parsear: {cell}. Error: {e}")
        return []


def update_counter_label():
    """Updates the counter label with the current file number."""
    counter_label.config(text=f"{current_file_index}/{total_files}")

def next_file():
    """Moves to the next file, if possible."""
    global current_file_index
    if current_file_index < total_files:
        current_file_index += 1
        update_counter_label()

def previous_file():
    """Moves to the previous file, if possible."""
    global current_file_index
    if current_file_index > 0:
        current_file_index -= 1
        update_counter_label()


def ask_user_name():
    """Asks the user to enter a name."""
    global user_name
    user_name = simpledialog.askstring("ID", "Enter user ID:")
    if not user_name:  # If the user doesn't enter anything
        messagebox.showerror("Error", "You must enter a name to continue.")
        ask_user_name()  # Ask for the name again


# Function to load the corresponding image from the selected options
def update_image_from_selection():
    global image_id1
    selected_label = selected_color.get()
    
    if not selected_label:
        shared_canvas.delete("image1")  # Borrar solo la imagen 1
        return
    
    directory = "Datasets/vita_tooth_test" if image_source.get() == "tooth" else "Datasets/reference_colors"
    new_size = (62,90) if directory == "Datasets/vita_tooth_test" else (50,50)
    image_path = os.path.join(directory, f"{selected_label}.png")
    
    if os.path.exists(image_path):
        img = Image.open(image_path).convert("RGBA").resize((new_size), Image.Resampling.LANCZOS)
        img_tk = ImageTk.PhotoImage(img)
        shared_canvas.delete("image1")  # Borrar cualquier imagen previa
        shared_canvas.image1 = img_tk  # Guardar referencia
        image_id1 = shared_canvas.create_image(*coords_img1, image=img_tk, tag="image1")

        shared_canvas.bind("<ButtonPress-1>", on_image_press)
        shared_canvas.bind("<B1-Motion>", on_image_drag)
        shared_canvas.bind("<ButtonRelease-1>", on_image_release)
    else:
        shared_canvas.delete("image1")
        shared_canvas.create_text(*coords_img1, text=" ", tag="image1")


def on_image_press(event):
    global offset_x, offset_y
    # Guardar las coordenadas iniciales del clic dentro del Canvas
    offset_x = event.x
    offset_y = event.y

def on_image_drag(event):
    global offset_x, offset_y
    # Mover la imagen izquierda dentro del Canvas si el clic fue sobre ella
    canvas_items = shared_canvas.find_closest(event.x, event.y)
    if canvas_items and canvas_items[0] == image_id1:
        dx = event.x - offset_x
        dy = event.y - offset_y
        shared_canvas.move(image_id1, dx, dy)
        offset_x = event.x
        offset_y = event.y

def on_image_release(event):
    global original_position_left_image, image_id1
    # Volver a la posición original de la imagen izquierda
    shared_canvas.coords(image_id1, original_position_left_image[0], original_position_left_image[1])


# Function to update images according to the selected Radiobutton
def update_image_source():
    update_image_from_selection()


def load_vita_images():
    """Loads all vita_tooth images and prepares them for random display."""
    global vita_images
    vita_images = []

    # Search for images in the vita_tooth directory
    for filename in os.listdir(vita_dir):
        if filename.endswith(".png"):
            image_path = os.path.join(vita_dir, filename)
            img = Image.open(image_path).convert("RGBA")  # Load with transparency
            img = img.resize((62, 90), Image.Resampling.LANCZOS)
            vita_images.append((filename, ImageTk.PhotoImage(img)))  # Save the name and image

    # Shuffle images at the start
    random.shuffle(vita_images)        


def show_next_image():
    """Shows the next random image in Image 2 and resets the values."""
    global current_index

    # Save results of the current case
    # if current_index >= 0:
    #     current_tooth = vita_images[current_index][0].split(".")[0]  # Get the tooth name
    #     update_results_matrix(current_tooth)

    # Move to the next index
    current_index += 1
    next_file()

    # If there are still images available
    if current_index < len(vita_images):  
        _, img_tk = vita_images[current_index]

        # Display the image on the canvas
        shared_canvas.delete("image2")  
        shared_canvas.image2 = img_tk  
        shared_canvas.create_image(*coords_img2, image=img_tk, tag="image2")

        # Restore saved values or reset inputs
        current_tooth = vita_images[current_index][0].split(".")[0] 
        current_idx_t = color_labels.index(current_tooth)
        current_row = data.iloc[current_idx_t]

        top_values = current_row['top']    
        middle_values = current_row['middle']
        bottom_values = current_row['bottom']

        # Para "top"
        all_static_texts[0].config(text=top_values[0][0])
        all_static_texts[1].config(text=top_values[1][0])
        all_static_texts[2].config(text=top_values[2][0])
        radiobutton_values[0].set(top_values[0][1])
        radiobutton_values[1].set(top_values[1][1])
        radiobutton_values[2].set(top_values[2][1])

        # Para "middle"
        all_static_texts[3].config(text=middle_values[0][0])
        all_static_texts[4].config(text=middle_values[1][0])
        all_static_texts[5].config(text=middle_values[2][0])
        radiobutton_values[3].set(middle_values[0][1])
        radiobutton_values[4].set(middle_values[1][1])
        radiobutton_values[5].set(middle_values[2][1])

        # Para "bottom"
        all_static_texts[6].config(text=bottom_values[0][0])
        all_static_texts[7].config(text=bottom_values[1][0])
        all_static_texts[8].config(text=bottom_values[2][0])
        radiobutton_values[6].set(bottom_values[0][1])
        radiobutton_values[7].set(bottom_values[1][1])
        radiobutton_values[8].set(bottom_values[2][1])

        
        # CUANDO VAYA GUARDANDO LA DATA AQUI COMPRUEBO SI GUARDADA
        # if current_row and any(value is not None for value in current_row):
        #     restore_previous_values(current_tooth)
        # else:
        #     reset_all_inputs()
        reset_all_inputs()

        # Update button states
        prev_button.config(state="normal" if current_index > 0 else "disabled")
        next_button.config(state="normal" if current_index < len(vita_images) - 1 else "disabled")
        update_image_from_selection()
        validate_first_column()  # Automatically validate if the first column is filled

    else:
        # All images have been shown, ask user if they want to finalize
        finalize = messagebox.askyesno("Finalize", "Your selections will be saved. Do you want to finalize?")
        if finalize:
            save_results_to_excel()
            prev_button.config(state="disabled")
            next_button.config(state="disabled")  # Disable "Next"
            reset_button.config(state="normal")  # Enable "Reset all"
        else:
            current_index -= 1  # Return to the last index


def show_previous_image():
    """Shows the previous image and restores its values."""
    global current_index

    if current_index > 0:  # Prevent going to negative indices
        # Save current results
        current_tooth = vita_images[current_index][0].split(".")[0]
        update_results_matrix(current_tooth)

        # Move to the previous index
        current_index -= 1

        # Display the previous image
        _, img_tk = vita_images[current_index]
        shared_canvas.delete("image2")
        shared_canvas.image2 = img_tk
        shared_canvas.create_image(*coords_img2, image=img_tk, tag="image2")

        # Restore saved values or reset inputs
        current_tooth = vita_images[current_index][0].split(".")[0] 
        current_idx_t = color_labels.index(current_tooth)
        current_row = data.iloc[current_idx_t]

        top_values = current_row['top']    
        middle_values = current_row['middle']
        bottom_values = current_row['bottom']

        # Para "top"
        all_static_texts[0].config(text=top_values[0][0])
        all_static_texts[1].config(text=top_values[1][0])
        all_static_texts[2].config(text=top_values[2][0])
        radiobutton_values[0].set(top_values[0][1])
        radiobutton_values[1].set(top_values[1][1])
        radiobutton_values[2].set(top_values[2][1])

        # Para "middle"
        all_static_texts[3].config(text=middle_values[0][0])
        all_static_texts[4].config(text=middle_values[1][0])
        all_static_texts[5].config(text=middle_values[2][0])
        radiobutton_values[3].set(middle_values[0][1])
        radiobutton_values[4].set(middle_values[1][1])
        radiobutton_values[5].set(middle_values[2][1])

        # Para "bottom"
        all_static_texts[6].config(text=bottom_values[0][0])
        all_static_texts[7].config(text=bottom_values[1][0])
        all_static_texts[8].config(text=bottom_values[2][0])
        radiobutton_values[6].set(bottom_values[0][1])
        radiobutton_values[7].set(bottom_values[1][1])
        radiobutton_values[8].set(bottom_values[2][1])


        update_image_from_selection()

    # Update button states
    prev_button.config(state="normal" if current_index > 0 else "disabled")
    next_button.config(state="normal" if current_index < len(vita_images) - 1 else "disabled")
    previous_file()


def restore_previous_values(current_tooth):
    """Restores the dropdown and slider values for the current index or resets inputs if no values exist."""
    global results_matrix

    try:
        # Retrieve the index for the given tooth
        current_index = color_labels.index(current_tooth)
    except ValueError:
        messagebox.showerror("Error", f"Tooth '{current_tooth}' not found in the list.")
        return

    # Retrieve the current row in the results matrix
    current_row = results_matrix[current_index]

    # Check if the current row has any non-None values
    if current_row and any(value is not None for value in current_row):
        # Extract saved values from the matrix row
        upper_value, upper_confidence, central_value, central_confidence, lower_value, lower_confidence = current_row

        # Parse the string values back to lists
        upper_value = upper_value.split(", ")
        upper_confidence = list(map(float, upper_confidence.split(", ")))
        central_value = central_value.split(", ")
        central_confidence = list(map(float, central_confidence.split(", ")))
        lower_value = lower_value.split(", ")
        lower_confidence = list(map(float, lower_confidence.split(", ")))

        # Restore the dropdowns and sliders
        # Para "top"
        all_static_texts[0].config(text=top_values[0][0])
        all_static_texts[1].config(text=top_values[1][0])
        all_static_texts[2].config(text=top_values[2][0])
        radiobutton_values[0].set(top_values[0][1])
        radiobutton_values[1].set(top_values[1][1])
        radiobutton_values[2].set(top_values[2][1])

        # Para "middle"
        all_static_texts[3].config(text=middle_values[0][0])
        all_static_texts[4].config(text=middle_values[1][0])
        all_static_texts[5].config(text=middle_values[2][0])
        radiobutton_values[3].set(middle_values[0][1])
        radiobutton_values[4].set(middle_values[1][1])
        radiobutton_values[5].set(middle_values[2][1])

        # Para "bottom"
        all_static_texts[6].config(text=bottom_values[0][0])
        all_static_texts[7].config(text=bottom_values[1][0])
        all_static_texts[8].config(text=bottom_values[2][0])
        radiobutton_values[6].set(bottom_values[0][1])
        radiobutton_values[7].set(bottom_values[1][1])
        radiobutton_values[8].set(bottom_values[2][1])

    else:
        # If the row is empty or contains only None, reset all inputs
        reset_all_inputs()



def validate_first_column():
    """Validates that the Sliders in the first column are filled."""
    all_valid = True

    # Validate that all Sliders have a value greater than 0
    # if all_valid:  # Only continue if the Comboboxes are valid
    #     for scale in additional_sliders:
    #         if scale.get() == 0.0:  # Slider at its initial position
    #             all_valid = False
    #             break

    # Enable or disable the "Next" button based on validation
    next_button.config(state="normal" if all_valid else "disabled")


def reset_all_inputs():
    """Resets all Comboboxes and Sliders."""
    # Reset all Sliders
    # for scale in additional_sliders:
    #     scale.set(3)  # Initial value


def reset_cycle():
    """Resets the image cycle and controls."""
    global current_index
    global user_name

    # Save results to Excel before resetting
    initialize_results_matrix()

    # Ask for the name again
    ask_user_name()

    # Shuffle images again
    random.shuffle(vita_images)  
    current_index = -1  # Reset index
    shared_canvas.delete("image2")  # Clear the second image canvas
    next_button.config(state="normal")  # Enable the "Next" button
    reset_button.config(state="disabled")  # Disable the "Reset all" button
    show_next_image()
    prev_button.config(state="disabled")


# Initialize results matrix with empty rows
def initialize_results_matrix():
    global results_matrix
    results_matrix = [[None] * 6 for _ in color_labels]  # 6 columns for each tooth


def update_results_matrix(diente):
    """Updates the results matrix with the values from dropdowns and sliders."""



def save_results_to_excel():
    """Saves the results matrix to an Excel file."""
    global user_name

    # File and sheet
    file_name = "Results.xlsx"
    sheet_name = user_name

    # Create file if it doesn't exist
    if not os.path.exists(file_name):
        wb = Workbook()
        wb.save(file_name)

    # Load the existing file
    wb = load_workbook(file_name)

    # If the sheet already exists, add an incremental suffix
    original_sheet_name = sheet_name
    counter = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{original_sheet_name}_{counter}"
        counter += 1

    # Create a new sheet
    ws = wb.create_sheet(title=sheet_name)

    # Add headers
    headers = [
        "Tooth",
        "Upper Value", "Upper Confidence",
        "Central Value", "Central Confidence",
        "Lower Value", "Lower Confidence"
    ]
    ws.append(headers)

    # Write the data from the matrix to Excel
    for tooth_label, row_data in zip(color_labels, results_matrix):
        ws.append([tooth_label] + row_data)

    # Save the file
    wb.save(file_name)
    messagebox.showinfo("Success", f"Results saved to {file_name}.")






############################################################ Main Windown ############################################################

root = tk.Tk()
# root.geometry("1300x800") 
root.attributes('-fullscreen', True)
root.bind('<Escape>', exit_fullscreen)
root.title("Color Selector")

# Excel with PyFCS data
file_path = os.path.join(os.getcwd(), "Datasets", "results_opt_1.xlsx") 
data = pd.read_excel(file_path)

columns_to_convert = ['top', 'middle', 'bottom']

# Convertir las cadenas a listas de tuplas para las columnas especificadas
for col in columns_to_convert:
    data[col] = data[col].apply(ast.literal_eval)

for idx, row in data.iterrows():
    top_values = row['top']    
    middle_values = row['middle']  
    bottom_values = row['bottom'] 


############################################################ INTERFACE ############################################################

# Create the main frame for the colors
frame = tk.Frame(root)
frame.pack(pady=10, padx=10)

# Global variable to store the user's name
user_name = ""
results_matrix = [] 
original_position_left_image = (150, 200)  # left image original position
left_image_id = None  # ID left image
current_file_index = 0

# Ask for the user's ID at the start
ask_user_name()

# Variable to store the selected option
image_source = tk.StringVar(value="tooth")  # By default, use 'Vita Tooth'

# Create the images and assign checkboxes
color_labels = ["A1", "A2", "A3", "A3_5", "A4", "B1", "B2", "B3", "B4", "C1", "C2", "C3", "C4", "D2", "D3", "D4"]
image_dir = os.path.join(os.getcwd(), "Datasets", "reference_colors")  # Initial image directory
vita_dir = os.path.join(os.getcwd(), "Datasets", "vita_tooth")  # Image directory for 'Image 2'
image_files = [f"{label}.png" for label in color_labels]

total_files = len([f for f in os.listdir(vita_dir) if os.path.isfile(os.path.join(vita_dir, f))])

# Create matrix to save results
initialize_results_matrix()

# Load images
images = [tk.PhotoImage(file=os.path.join(image_dir, image)) for image in image_files]

# Shared variable for the checkboxes
selected_color = tk.StringVar(value=0)  

# Create containers for images and checkboxes
for idx, (color, img) in enumerate(zip(color_labels, images)):
    col_frame = tk.Frame(frame)  # Frame for each image and checkbox
    col_frame.grid(row=0, column=idx, padx=5, pady=5)

    # Image
    img_label = tk.Label(col_frame, image=img)
    img_label.pack()

    # Radiobutton under the image
    chk = tk.Radiobutton(
        col_frame, 
        text=color, 
        variable=selected_color, 
        value=color, 
        command=update_image_from_selection  
    )
    chk.pack()

# Create a main frame for organizing the center and right sections
main_center_frame = tk.Frame(root)
main_center_frame.pack(pady=20, padx=20)

# Create the area for central images
center_frame = tk.Frame(main_center_frame)
center_frame.grid(row=0, column=0, padx=10)

# Reduce top spacers to bring down the images
tk.Label(center_frame, height=0).grid(row=0, column=0, columnspan=2)  

# Create canvas for two images
shared_canvas = tk.Canvas(center_frame, width=440, height=400, bg="SystemButtonFace", highlightthickness=0)
shared_canvas.grid(row=1, column=0, padx=10, pady=1, columnspan=2)

# Coordenates of the two images 
coords_img1 = original_position_left_image
coords_img2 = (300, 200)  

# Buttons and checkboxes under the central images
buttons_frame = tk.Frame(center_frame)
buttons_frame.grid(row=3, column=0, columnspan=2, pady=2)

counter_label = tk.Label(buttons_frame, text="1/16", font=("Arial", 10))
counter_label.grid(row=0, column=0, columnspan=2, pady=5)

prev_button = tk.Button(buttons_frame, text="Previous", command=show_previous_image, state="disabled")
prev_button.grid(row=1, column=0, padx=10)

next_button = tk.Button(buttons_frame, text="Next", command=show_next_image)
next_button.grid(row=1, column=1, padx=10)

# "Reset all" button to restart the cycle
reset_button = tk.Button(buttons_frame, text="Reset all", command=reset_cycle, state="disabled")
reset_button.grid(row=3, column=0, columnspan=2, pady=5)

# Create the sliders with comboboxes to the right of the central images
sliders_frame = tk.Frame(main_center_frame)
sliders_frame.grid(row=0, column=1, padx=10, pady=1)

# Create global lists 
all_static_texts = []  # Referencias a textos estáticos
all_radiobuttons = []  # Referencias a radiobuttons
radiobutton_values = []  # Almacenará los valores seleccionados de los radiobuttons

# Nombres de filas
row_names = ["Upper Tooth", "Central Tooth", "Lower Tooth"]

for row_idx, row_name in enumerate(row_names):
    # Etiqueta para cada fila
    row_label = tk.Label(sliders_frame, text=row_name, font=("Arial", 10, "bold"))
    row_label.grid(row=row_idx, column=0, padx=10, pady=0)

    # Agregar texto estático y radiobuttons
    for col_idx in range(3):  # Tres columnas por fila
        control_frame = tk.Frame(sliders_frame)
        control_frame.grid(row=row_idx, column=col_idx + 1, padx=10, pady=7)

        # Texto estático
        static_text = tk.Label(control_frame, text="Text", font=("Arial", 10))
        static_text.pack()
        all_static_texts.append(static_text)  # Guardar referencia

        # Radiobuttons
        radiobutton_frame = tk.Frame(control_frame)
        radiobutton_frame.pack()

        rb_value = tk.IntVar(value=3)  # Valor predeterminado de los radiobuttons
        radiobutton_values.append(rb_value)

        for rb_idx in range(1, 6):  # Crear 5 radiobuttons (valores 1-5)
            rb = tk.Radiobutton(
                radiobutton_frame, 
                text=str(rb_idx),      # El texto del radiobutton
                variable=rb_value, 
                value=rb_idx,
                indicatoron=False,     # Hace que el botón sea cuadrado con texto alineado
                width=2                # Ajusta el ancho para mejorar la apariencia
            )
            rb.pack(side=tk.LEFT)
            all_radiobuttons.append(rb)

# Create a new frame for the additional sliders to the right
additional_text_frame = tk.Frame(main_center_frame)
additional_text_frame.grid(row=0, column=2, padx=10, pady=1)

# Agregar un encabezado para la nueva columna
header_label = tk.Label(additional_text_frame, text="Comments", font=("Arial", 12, "bold"))
header_label.grid(row=0, column=0, padx=0, pady=0, columnspan=2)

# Crear una lista para almacenar las referencias de las casillas de texto
additional_text_entries = []

# Agregar casillas de texto para cada fila
for row_idx in range(len(row_names)):
    # Crear una casilla de texto
    text_entry = tk.Entry(additional_text_frame, width=15, font=("Arial", 10))
    text_entry.grid(row=row_idx + 1, column=0, padx=10, pady=7)  # Espaciado para cada fila

    # Guardar la referencia a la casilla de texto
    additional_text_entries.append(text_entry)

# Disable the "Next" button at the start
next_button.config(state="disabled")

vita_images = []
current_index = -1


# Create Radiobuttons to select the reference type
reference_options_frame = tk.Frame(buttons_frame)
reference_options_frame.grid(row=2, column=0, columnspan=2, pady=10)

# When selected, image_source will be "tooth"
tk.Radiobutton(
    reference_options_frame,
    text="Reference Tooth",
    variable=image_source,
    value="tooth",  
    command=update_image_source  
).grid(row=0, column=0, padx=10)

# When selected, image_source will be "color"
tk.Radiobutton(
    reference_options_frame,
    text="Reference Color",
    variable=image_source,
    value="color",  
    command=update_image_source  
).grid(row=0, column=1, padx=10)

# Load the images at the start
load_vita_images()
update_image_from_selection()

# Show the first random image
show_next_image()
prev_button.config(state="disabled")

# Run the interface
root.mainloop()
