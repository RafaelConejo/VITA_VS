import os
import sys
import random
import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import ttk, simpledialog, messagebox
from PIL import Image, ImageTk

current_dir = os.path.dirname(__file__)
sys.path.append(current_dir)


############################################################ Funcions ############################################################

def exit_fullscreen(event):
    root.attributes('-fullscreen', False)


def ask_user_name():
    """Asks the user to enter a name."""
    global user_name
    user_name = simpledialog.askstring("ID", "Enter user ID:")
    if not user_name:  # If the user doesn't enter anything
        messagebox.showerror("Error", "You must enter a name to continue.")
        ask_user_name()  # Ask for the name again


# Function to load the corresponding image from the selected options
def update_image_from_selection():
    global image_id1
    selected_label = selected_color.get()
    
    if not selected_label:
        shared_canvas.delete("image1")  # Borrar solo la imagen 1
        return
    
    directory = "Datasets/vita_tooth_test" if image_source.get() == "tooth" else "Datasets/reference_colors"
    new_size = (62,90) if directory == "Datasets/vita_tooth_test" else (50,50)
    image_path = os.path.join(directory, f"{selected_label}.png")
    
    if os.path.exists(image_path):
        img = Image.open(image_path).convert("RGBA").resize((new_size), Image.Resampling.LANCZOS)
        img_tk = ImageTk.PhotoImage(img)
        shared_canvas.delete("image1")  # Borrar cualquier imagen previa
        shared_canvas.image1 = img_tk  # Guardar referencia
        image_id1 = shared_canvas.create_image(*coords_img1, image=img_tk, tag="image1")

        shared_canvas.bind("<ButtonPress-1>", on_image_press)
        shared_canvas.bind("<B1-Motion>", on_image_drag)
        shared_canvas.bind("<ButtonRelease-1>", on_image_release)
    else:
        shared_canvas.delete("image1")
        shared_canvas.create_text(*coords_img1, text=" ", tag="image1")


def on_image_press(event):
    global offset_x, offset_y
    # Guardar las coordenadas iniciales del clic dentro del Canvas
    offset_x = event.x
    offset_y = event.y

def on_image_drag(event):
    global offset_x, offset_y
    # Mover la imagen izquierda dentro del Canvas si el clic fue sobre ella
    canvas_items = shared_canvas.find_closest(event.x, event.y)
    if canvas_items and canvas_items[0] == image_id1:
        dx = event.x - offset_x
        dy = event.y - offset_y
        shared_canvas.move(image_id1, dx, dy)
        offset_x = event.x
        offset_y = event.y

def on_image_release(event):
    global original_position_left_image, image_id1
    # Volver a la posición original de la imagen izquierda
    shared_canvas.coords(image_id1, original_position_left_image[0], original_position_left_image[1])


# Function to update images according to the selected Radiobutton
def update_image_source():
    update_image_from_selection()


def load_vita_images():
    """Loads all vita_tooth images and prepares them for random display."""
    global vita_images
    vita_images = []

    # Search for images in the vita_tooth directory
    for filename in os.listdir(vita_dir):
        if filename.endswith(".png"):
            image_path = os.path.join(vita_dir, filename)
            img = Image.open(image_path).convert("RGBA")  # Load with transparency
            img = img.resize((62, 90), Image.Resampling.LANCZOS)
            vita_images.append((filename, ImageTk.PhotoImage(img)))  # Save the name and image

    # Shuffle images at the start
    random.shuffle(vita_images)        


def show_next_image():
    """Shows the next random image in Image 2 and resets the values."""
    global current_index

    # Save results of the current case
    if current_index >= 0:
        current_tooth = vita_images[current_index][0].split(".")[0]  # Get the tooth name
        update_results_matrix(current_tooth)

    # Move to the next index
    current_index += 1

    if current_index < len(vita_images):  # If there are still images available
        _, img_tk = vita_images[current_index]

        # Display the image on the canvas
        shared_canvas.delete("image2")  # Borrar cualquier imagen previa
        shared_canvas.image2 = img_tk  # Guardar referencia
        shared_canvas.create_image(*coords_img2, image=img_tk, tag="image2")

        # Reset all Comboboxes and Sliders
        reset_all_inputs()

        # Disable the "Next" button if the first column is not filled
        next_button.config(state="disabled")
        validate_first_column()  # Automatically validate if the first column is filled

    else:  
        # All images have been shown
        save_results_to_excel()
        next_button.config(state="disabled")  # Disable "Next"
        reset_button.config(state="normal")  # Enable "Reset all"


def validate_first_column():
    """Validates that the Comboboxes and Sliders in the first column are filled."""
    all_valid = True

    # Validate that all Comboboxes have a value different from the initial
    for combo in first_column_comboboxes:
        if combo.get() == " ":  # Default value
            all_valid = False
            break

    # Validate that all Sliders have a value greater than 0
    if all_valid:  # Only continue if the Comboboxes are valid
        for scale in first_column_scales:
            if scale.get() == 0.0:  # Slider at its initial position
                all_valid = False
                break

    # Enable or disable the "Next" button based on validation
    next_button.config(state="normal" if all_valid else "disabled")


def reset_all_inputs():
    """Resets all Comboboxes and Sliders."""
    # Reset all Comboboxes
    for combo in all_comboboxes:
        combo.set(" ")  # Initial value

    # Reset all Sliders
    for scale in all_scales:
        scale.set(0)  # Initial value


def reset_cycle():
    """Resets the image cycle and controls."""
    global current_index
    global user_name

    # Save results to Excel before resetting
    initialize_results_matrix()

    # Ask for the name again
    ask_user_name()

    # Shuffle images again
    random.shuffle(vita_images)  
    current_index = -1  # Reset index
    shared_canvas.delete("image2")  # Clear the second image canvas
    next_button.config(state="normal")  # Enable the "Next" button
    reset_button.config(state="disabled")  # Disable the "Reset all" button
    show_next_image()


# Initialize results matrix with empty rows
def initialize_results_matrix():
    global results_matrix
    results_matrix = [[None] * 6 for _ in color_labels]  # 6 columns for each tooth


def update_results_matrix(diente):
    """Updates the results matrix with the values from dropdowns and sliders."""
    global results_matrix

    # Find the row corresponding to the current tooth
    try:
        row_index = color_labels.index(diente)
    except ValueError:
        messagebox.showerror("Error", f"Tooth '{diente}' not found in the list.")
        return

    # Extract values from the dropdowns and sliders
    upper_value = [all_comboboxes[0].get(), 
                   all_comboboxes[1].get(), 
                   all_comboboxes[2].get()]

    upper_confidence = [all_scales[0].get(), 
                        all_scales[1].get(), 
                        all_scales[2].get()]

    central_value = [all_comboboxes[3].get(), 
                     all_comboboxes[4].get(), 
                     all_comboboxes[5].get()]

    central_confidence = [all_scales[3].get(), 
                          all_scales[4].get(), 
                          all_scales[5].get()]

    lower_value = [all_comboboxes[6].get(), 
                   all_comboboxes[7].get(), 
                   all_comboboxes[8].get()]

    lower_confidence = [all_scales[6].get(), 
                        all_scales[7].get(), 
                        all_scales[8].get()]

    # Update the corresponding row
    results_matrix[row_index] = [
    ', '.join(map(str, upper_value)),
    ', '.join(map(str, upper_confidence)),
    ', '.join(map(str, central_value)),
    ', '.join(map(str, central_confidence)),
    ', '.join(map(str, lower_value)),
    ', '.join(map(str, lower_confidence))
]


def save_results_to_excel():
    """Saves the results matrix to an Excel file."""
    global user_name

    # File and sheet
    file_name = "Results.xlsx"
    sheet_name = user_name

    # Create file if it doesn't exist
    if not os.path.exists(file_name):
        wb = Workbook()
        wb.save(file_name)

    # Load the existing file
    wb = load_workbook(file_name)

    # If the sheet already exists, add an incremental suffix
    original_sheet_name = sheet_name
    counter = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{original_sheet_name}_{counter}"
        counter += 1

    # Create a new sheet
    ws = wb.create_sheet(title=sheet_name)

    # Add headers
    headers = [
        "Tooth",
        "Upper Value", "Upper Confidence",
        "Central Value", "Central Confidence",
        "Lower Value", "Lower Confidence"
    ]
    ws.append(headers)

    # Write the data from the matrix to Excel
    for tooth_label, row_data in zip(color_labels, results_matrix):
        ws.append([tooth_label] + row_data)

    # Save the file
    wb.save(file_name)
    messagebox.showinfo("Success", f"Results saved to {file_name}.")






############################################################ Main Windown ############################################################

root = tk.Tk()
# root.geometry("1300x800") 
root.attributes('-fullscreen', True)
root.bind('<Escape>', exit_fullscreen)
root.title("Color Selector")


############################################################ INTERFACE ############################################################

# Create the main frame for the colors
frame = tk.Frame(root)
frame.pack(pady=10, padx=10)

# Global variable to store the user's name
user_name = ""
results_matrix = [] 
original_position_left_image = (150, 200)  # left image original position
left_image_id = None  # ID left image

# Ask for the user's ID at the start
ask_user_name()

# Variable to store the selected option
image_source = tk.StringVar(value="tooth")  # By default, use 'Vita Tooth'

# Create the images and assign checkboxes
color_labels = ["A1", "A2", "A3", "A3_5", "A4", "B1", "B2", "B3", "B4", "C1", "C2", "C3", "C4", "D2", "D3", "D4"]
image_dir = os.path.join(os.getcwd(), "Datasets", "reference_colors")  # Initial image directory
vita_dir = os.path.join(os.getcwd(), "Datasets", "vita_tooth")  # Image directory for 'Image 2'
image_files = [f"{label}.png" for label in color_labels]

# Create matrix to save results
initialize_results_matrix()

# Load images
images = [tk.PhotoImage(file=os.path.join(image_dir, image)) for image in image_files]

# Shared variable for the checkboxes
selected_color = tk.StringVar(value=0)  

# Create containers for images and checkboxes
for idx, (color, img) in enumerate(zip(color_labels, images)):
    col_frame = tk.Frame(frame)  # Frame for each image and checkbox
    col_frame.grid(row=0, column=idx, padx=5, pady=5)

    # Image
    img_label = tk.Label(col_frame, image=img)
    img_label.pack()

    # Radiobutton under the image
    chk = tk.Radiobutton(
        col_frame, 
        text=color, 
        variable=selected_color, 
        value=color, 
        command=update_image_from_selection  
    )
    chk.pack()

# Create a main frame for organizing the center and right sections
main_center_frame = tk.Frame(root)
main_center_frame.pack(pady=20, padx=20)

# Create the area for central images
center_frame = tk.Frame(main_center_frame)
center_frame.grid(row=0, column=0, padx=10)

# Reduce top spacers to bring down the images
tk.Label(center_frame, height=0).grid(row=0, column=0, columnspan=2)  

# Create canvas for two images
shared_canvas = tk.Canvas(center_frame, width=440, height=400, bg="SystemButtonFace", highlightthickness=0)
shared_canvas.grid(row=1, column=0, padx=10, pady=1, columnspan=2)

# Coordenates of the two images 
coords_img1 = original_position_left_image
coords_img2 = (300, 200)  

# Buttons and checkboxes under the central images
buttons_frame = tk.Frame(center_frame)
buttons_frame.grid(row=2, column=0, columnspan=2, pady=2)

next_button = tk.Button(buttons_frame, text="Next", command=show_next_image)
next_button.grid(row=0, column=0, columnspan=2, pady=5)

# "Reset all" button to restart the cycle
reset_button = tk.Button(buttons_frame, text="Reset all", command=reset_cycle, state="disabled")
reset_button.grid(row=2, column=0, columnspan=2, pady=5)

# Create the sliders with comboboxes to the right of the central images
sliders_frame = tk.Frame(main_center_frame)
sliders_frame.grid(row=0, column=1, padx=10, pady=1)

# Create global lists to store references to the Comboboxes and Sliders
all_comboboxes = []  # All Comboboxes
all_scales = []      # All Sliders

# Create specific lists for the first column (for validation)
first_column_comboboxes = []
first_column_scales = []

# Row names
row_names = ["Upper Tooth", "Central Tooth", "Lower Tooth"]
for row_idx, row_name in enumerate(row_names):
    # Label for each row
    row_label = tk.Label(sliders_frame, text=row_name, font=("Arial", 10, "bold"))
    row_label.grid(row=row_idx, column=0, padx=10, pady=0) 

    # Add sliders and comboboxes in corresponding columns
    for col_idx in range(3):  # Three columns of sliders and comboboxes per row
        slider_frame = tk.Frame(sliders_frame)
        slider_frame.grid(row=row_idx, column=col_idx + 1, padx=10, pady=7)

        # Combobox
        combo = ttk.Combobox(slider_frame, values=color_labels, width=4, state="readonly")
        combo.set(" ")  # Initial value
        combo.pack()
        all_comboboxes.append(combo)  # Save reference globally

        if col_idx == 0:  # First column
            first_column_comboboxes.append(combo)

        # Slider
        scale = tk.Scale(slider_frame, from_=0, to=1, resolution=0.1, orient=tk.HORIZONTAL)
        scale.pack()
        all_scales.append(scale)  # Save reference globally

        if col_idx == 0:  # First column
            first_column_scales.append(scale)

# Disable the "Next" button at the start
next_button.config(state="disabled")

# Bind validation to events
for combo in first_column_comboboxes:
    combo.bind("<<ComboboxSelected>>", lambda e: validate_first_column())

for scale in first_column_scales:
    scale.bind("<ButtonRelease-1>", lambda e: validate_first_column())
    scale.bind("<Motion>", lambda e: validate_first_column())

vita_images = []
current_index = -1


# Create Radiobuttons to select the reference type
reference_options_frame = tk.Frame(buttons_frame)
reference_options_frame.grid(row=1, column=0, columnspan=2, pady=10)

# When selected, image_source will be "tooth"
tk.Radiobutton(
    reference_options_frame,
    text="Reference Tooth",
    variable=image_source,
    value="tooth",  
    command=update_image_source  
).grid(row=0, column=0, padx=10)

# When selected, image_source will be "color"
tk.Radiobutton(
    reference_options_frame,
    text="Reference Color",
    variable=image_source,
    value="color",  
    command=update_image_source  
).grid(row=0, column=1, padx=10)

# Load the images at the start
load_vita_images()
update_image_from_selection()

# Show the first random image
show_next_image()

# Run the interface
root.mainloop()